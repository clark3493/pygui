import numpy as np
import tkinter as tk

from openpyxl.utils import get_column_letter
from tkinter import Scrollbar


def cellname(j, i, style='excel'):
    if style == 'excel':
        return f'{get_column_letter(i)}{str(j)}'
    elif style == 'array':
        return f'{j},{i}'
    else:
        raise ValueError(f"Unrecognized cell name style: {style}")


class Cell(tk.Entry):

    def __init__(self, parent, j, i, siblings, **kwargs):
        self.var = tk.StringVar()
        super().__init__(parent, textvariable=self.var)

        self.col = i
        self.row = j
        self.siblings = siblings
        style = 'excel' if 'index_style' not in kwargs else kwargs['index_style']
        self.name = cellname(j, i, style=style)

        self.var.set('')

        self.options = _CellOptions(self, **kwargs)

    @property
    def readonly(self):
        return self.config()['state'] == tk.DISABLED

    @property
    def value(self):
        return self.var.get()

    @value.setter
    def value(self, value):
        self.var.set(value)


class _CellOptions(object):

    def __init__(self, cell, justify='right', readonly=True, index_style='excel'):
        self._justify  = None
        self._readonly = None
        self._index_style    = index_style

        self.cell      = cell

        self.justify   = justify
        self.readonly  = readonly

    @property
    def index_style(self):
        return self._index_style

    @property
    def justify(self):
        return self._justify

    @justify.setter
    def justify(self, value):
        self._justify = value
        self.cell.config(justify=value)

    @property
    def readonly(self):
        return self._readonly

    @readonly.setter
    def readonly(self, value):
        self._readonly = value
        state = tk.DISABLED if value is True else tk.NORMAL
        self.cell.config(state=state)


class Table(tk.Frame):

    def __init__(self, parent, nrows=0, ncols=0, data=None, **kwargs):
        super().__init__(parent)

        if data is not None:
            data = self._prep_data(data)
            nrows = nrows if nrows >= data.shape[0] else data.shape[0]
            ncols = ncols if ncols >= data.shape[1] else data.shape[1]

        self.nrows = nrows
        self.ncols = ncols
        self.cells = {}
        self._row_labels = []
        self._col_labels = []

        self.options = _TableOptions(self, **kwargs)

        self.pack()
        self._initialize_cells()
        #self._add_scrollbars()
        #self.create_window((0,0), window=self.cellframe, anchor='nw')

        if data is not None:
            self.set_data(data)

    @property
    def col_indices(self):
        if self.options.index_style == 'excel':
            return range(1, self.ncols+1)
        elif self.options.index_style == 'array':
            return range(self.ncols)

    @property
    def row_indices(self):
        if self.options.index_style == 'excel':
            return range(1, self.nrows+1)
        elif self.options.index_style == 'array':
            return range(self.nrows)

    def hide_labels(self):
        self.hide_row_labels()
        self.hide_col_labels()

    def hide_col_labels(self):
        for label in self._col_labels:
            label.grid_remove()

    def hide_row_labels(self):
        for label in self._row_labels:
            label.grid_remove()

    def show_labels(self):
        self.show_row_labels()
        self.show_col_labels()

    def show_row_labels(self):
        for label in self._row_labels:
            label.grid()

    def show_col_labels(self):
        for label in self._col_labels:
            label.grid()

    def set_data(self, data):
        data = self._prep_data(data)
        for j, row in enumerate(self.row_indices):
            for i, col in enumerate(self.col_indices):
                name = cellname(row, col, style=self.options.index_style)
                cell = self.cells[name]
                cell.value = data[j, i]
    """
    def _add_scrollbars(self):
        self.horizontal_scrollbar = Scrollbar(self, orient=tk.HORIZONTAL)
        self.vertical_scrollbar = Scrollbar(self, orient=tk.VERTICAL)

        self.horizontal_scrollbar.pack(fill=tk.X, side=tk.BOTTOM)
        self.vertical_scrollbar.pack(fill=tk.Y, side=tk.RIGHT)

        self['xscrollcommand'] = self.horizontal_scrollbar.set
        self['yscrollcommand'] = self.vertical_scrollbar.set

        self.horizontal_scrollbar['command'] = self.xview
        self.vertical_scrollbar['command'] = self.yview
    """
    def _initialize_cells(self):
        # Frame for all the cells
        self.cellframe = tk.Frame(self)
        self.cellframe.pack(side=tk.TOP, fill=tk.BOTH)

        # column_labels
        blank = tk.Label(self.cellframe)
        blank.grid(row=0, column=0)
        self._row_labels.append(blank)
        self._col_labels.append(blank)
        for i in range(1, self.ncols+1):
            label_text = str(i-1) if self.options.index_style == 'array' else get_column_letter(i)
            label = tk.Label(self.cellframe, text=label_text, font='bold')
            label.grid(row=0, column=i)
            self._col_labels.append(label)

        # fill in the rows
        for j in range(1, self.nrows+1):
            label_text = str(j-1) if self.options.index_style == 'array' else str(j)
            rowlabel = tk.Label(self.cellframe, text=label_text, font='bold')
            rowlabel.grid(row=j, column=0)
            self._row_labels.append(rowlabel)
            for i in range(1, self.ncols+1):
                row_index = j if self.options.index_style == 'excel' else j-1
                col_index = i if self.options.index_style == 'excel' else i-1
                cell = Cell(self.cellframe, row_index, col_index, self.cells,
                            readonly=self.options.readonly,
                            index_style=self.options.index_style)
                self.cells[cell.name] = cell
                cell.grid(row=j, column=i)

    @staticmethod
    def _prep_data(data):
        data = np.array(data)
        if len(data.shape) > 2:
            raise ValueError("Array of greater than 2-dimensions are not currently supported")
        if len(data.shape) == 1:
            data = data.reshape((1, data.shape[0]))
        return data


class _TableOptions(object):

    def __init__(self, table, readonly=True, index_style='excel', labels_on=True, **kwargs):
        self.table = table

        self._labels_on   = None
        self._readonly    = None
        self._index_style = index_style

        self.labels_on    = labels_on
        self.readonly     = readonly

    @property
    def index_style(self):
        return self._index_style

    @property
    def labels_on(self):
        return self._labels_on

    @labels_on.setter
    def labels_on(self, value):
        self._labels_on = value
        if value:
            self.table.show_labels()
        else:
            self.table.hide_labels()

    @property
    def readonly(self):
        return self._readonly

    @readonly.setter
    def readonly(self, value):
        self._readonly = value
        for cell in self.table.cells.values():
            cell.options.readonly = value


if __name__ == "__main__":
    root = tk.Tk()
    data = [[1., 2., 3.], [4., 5., 6.], [7., 8., 9.], [10., 11., 12.]]
    #app = Table(root, 5, 5, index_style='array')
    app = Table(root, data=data, index_style='array')
    #app.hide_labels()
    app.mainloop()
